import os
import json
import csv
from flask import Flask, render_template, request, redirect, url_for, jsonify
import io
from flask import send_file
# Custom modules
from config import CONFIG
from solver import generate_timetable, generate_timetable_with_retry
from adapter import build_solver_inputs_from_classes
from extractor import get_solver_data_from_pdf 

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = os.path.abspath('uploads')
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# ── Clear stale session files on every app startup ───────────────────────────
for _f in ["generated_timetable.json", "generated_metadata.json",
           "final_schedule.json", "temp_web_data.json",
           "last_extraction.json", "solver_input_debug.json"]:
    try:
        if os.path.exists(_f):
            os.remove(_f)
    except Exception:
        pass


@app.route("/")
def home():
    return render_template("upload.html")

@app.route("/upload-pdf", methods=["POST"])
def upload_pdf():
    file = request.files.get('file')
    if not file or file.filename == '':
        return "No file selected", 400

    pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], "uploaded_schedule.pdf")
    file.save(pdf_path)
    
    try:
        raw_data = get_solver_data_from_pdf(pdf_path) 
        with open("last_extraction.json", "w") as f:
            json.dump(raw_data, f)
        
        CONFIG["raw_extraction"] = raw_data
        return redirect(url_for('generate'))
        
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return f"AI Extraction Failed: {str(e)}", 500

@app.route("/generate")
def generate():
    data = CONFIG.get("raw_extraction")
    if not data and os.path.exists("last_extraction.json"):
        with open("last_extraction.json", "r") as f:
            data = json.load(f)
    
    if not data:
        return "<h3>No data found. Please upload a PDF first.</h3>"

    try:
        display_data = []
        teacher_map = data.get('teacher_list', {})

        # 1. Process Theory (Now a LIST, not a DICT)
        for class_id, teachers_list in data.get('class_teacher_periods', {}).items():
            # FIX: Loop through the list directly instead of using .items()
            for item in teachers_list:
                t_id = str(item.get('teacher_id'))
                subj = item.get('subject', 'Theory')
                p_val = item.get('periods', 0)

                display_data.append({
                    "class": f"Class {class_id}",
                    "subject": subj,
                    "teacher": teacher_map.get(t_id, {}).get('Name', f"S{t_id}"),
                    "type": "Theory",
                    "periods": p_val
                })

        # 2. Process Labs (Now a LIST, not a DICT)
        for class_id, labs_list in data.get('lab_teacher_periods', {}).items():
            # FIX: Loop through the list directly instead of using .items()
            for item in labs_list:
                t_id = str(item.get('teacher_id'))
                subj = item.get('subject', 'Lab')
                p_raw = item.get('periods', [0])
                p_count = p_raw[0] if isinstance(p_raw, list) else p_raw

                display_data.append({
                    "class": f"Class {class_id}",
                    "subject": subj,
                    "teacher": teacher_map.get(t_id, {}).get('Name', f"S{t_id}"),
                    "type": "Lab",
                    "periods": p_count
                })

        return render_template("view_simple.html", rows=display_data)

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return f"<h3>Data Processing Error: {str(e)}</h3>"
















# ─────────────────────────────────────────────────────────────────────────────
#  HELPER — resolve one cell value from the timetable
#  Timetable structure: timetable[slot_index][class_idx]
#  where slot_index = day * periods_per_day + period
# ─────────────────────────────────────────────────────────────────────────────
def _cell_text(timetable, class_idx, day, period, periods_per_day):
    slot_index = day * periods_per_day + period
    try:
        slot_row = timetable[slot_index]
        # slot_row is either a list [cls0_val, cls1_val, ...] or a dict
        if isinstance(slot_row, list):
            raw = slot_row[class_idx]
        elif isinstance(slot_row, dict):
            raw = slot_row.get(str(class_idx), slot_row.get(class_idx, ""))
        else:
            raw = slot_row
    except (IndexError, KeyError, TypeError):
        return "", "normal"

    if raw is None or raw == 0 or raw == "0":
        return "", "normal"

    text = str(raw).strip()
    lower = text.lower()

    if lower in ("free", "f") or (lower.startswith("f") and len(text) < 4):
        return "Free", "free"
    if "lab" in lower:
        return text, "lab"
    return text, "normal"


# ─────────────────────────────────────────────────────────────────────────────
#  EXCEL DOWNLOAD
# ─────────────────────────────────────────────────────────────────────────────
@app.route("/download/excel")
def download_excel():
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    if not os.path.exists("generated_timetable.json") or \
       not os.path.exists("generated_metadata.json") or \
       not os.path.exists("temp_web_data.json"):
        return "No timetable found. Please generate one first.", 404

    with open("generated_timetable.json")  as f: timetable   = json.load(f)
    with open("generated_metadata.json")   as f: meta        = json.load(f)
    with open("temp_web_data.json")        as f: stored      = json.load(f)

    days        = meta["days"]
    periods     = meta["periods"]
    num_classes = meta["num_classes"]

    organized_keys = list(stored.get("organized", {}).keys())
    all_class_names = [organized_keys[i] if i < len(organized_keys) else str(i+1) for i in range(num_classes)]

    # Support single-class export via ?class_idx=N
    single_idx = request.args.get("class_idx", None)
    if single_idx is not None:
        try:
            single_idx = int(single_idx)
            class_names = [all_class_names[single_idx]]
            class_indices = [single_idx]
        except (ValueError, IndexError):
            class_names = all_class_names
            class_indices = list(range(num_classes))
    else:
        class_names   = all_class_names
        class_indices = list(range(num_classes))

    # Day labels: Mon–Sat for 6, Mon–Fri for 5, etc.
    _day_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    day_labels = [_day_names[i] if i < len(_day_names) else f"Day {i+1}" for i in range(days)]

    # ── Style factories (new object per cell avoids openpyxl shared-style bugs) ─
    def hdr_fill():  return PatternFill("solid", fgColor="1F3864")
    def hdr_font():  return Font(color="FFFFFF", bold=True, size=11)
    def per_fill():  return PatternFill("solid", fgColor="D9E1F2")
    def per_font():  return Font(bold=True, size=10)
    def free_fill(): return PatternFill("solid", fgColor="FFF9C4")
    def lab_fill():  return PatternFill("solid", fgColor="E7F5FF")
    def norm_fill(): return PatternFill("solid", fgColor="FFFFFF")
    def mk_border(): return Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"))
    def mk_center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for cls_idx, cls_name in zip(class_indices, class_names):
        ws = wb.create_sheet(title=f"Class {cls_name}"[:31])

        # Header row: Day | P1 | P2 | P3 | ...
        ws.row_dimensions[1].height = 26
        ws.column_dimensions["A"].width = 14

        for col, label in enumerate(["Day"] + [f"P{p+1}" for p in range(periods)]):
            c = ws.cell(row=1, column=col+1, value=label)
            c.fill      = hdr_fill()
            c.font      = hdr_font()
            c.alignment = mk_center()
            c.border    = mk_border()

        col_letters = list("BCDEFGHIJKLMNOPQRSTUVWXYZ")
        for p in range(periods):
            if p < len(col_letters):
                ws.column_dimensions[col_letters[p]].width = 24

        # Data rows — one row per day
        for d in range(days):
            row_num = d + 2
            ws.row_dimensions[row_num].height = 42

            # Day label cell
            dc = ws.cell(row=row_num, column=1, value=day_labels[d])
            dc.fill      = per_fill()
            dc.font      = per_font()
            dc.alignment = mk_center()
            dc.border    = mk_border()

            for p in range(periods):
                text, kind = _cell_text(timetable, cls_idx, d, p, periods)

                fill = {"free": free_fill(), "lab": lab_fill()}.get(kind, norm_fill())

                pc = ws.cell(row=row_num, column=p+2, value=text)
                pc.fill      = fill
                pc.alignment = mk_center()
                pc.border    = mk_border()
                pc.font      = Font(size=9, bold=(kind == "lab"),
                                    italic=(kind == "free"),
                                    color="6C757D" if kind == "free" else "000000")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    fname = f"timetable_class_{class_names[0]}.xlsx" if len(class_names) == 1 else "timetable.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=fname
    )


# ─────────────────────────────────────────────────────────────────────────────
#  PDF DOWNLOAD
# ─────────────────────────────────────────────────────────────────────────────
@app.route("/download/pdf")
def download_pdf():
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    if not os.path.exists("generated_timetable.json") or \
       not os.path.exists("generated_metadata.json") or \
       not os.path.exists("temp_web_data.json"):
        return "No timetable found. Please generate one first.", 404

    with open("generated_timetable.json")  as f: timetable   = json.load(f)
    with open("generated_metadata.json")   as f: meta        = json.load(f)
    with open("temp_web_data.json")        as f: stored      = json.load(f)

    days        = meta["days"]
    periods     = meta["periods"]
    num_classes = meta["num_classes"]

    organized_keys = list(stored.get("organized", {}).keys())
    class_names = []
    for i in range(num_classes):
        class_names.append(organized_keys[i] if i < len(organized_keys) else str(i + 1))

    # Day labels: Mon–Sat for 6, Mon–Fri for 5, etc.
    _day_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    day_labels = [_day_names[i] if i < len(_day_names) else f"Day {i+1}" for i in range(days)]

    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=1.5*cm,  bottomMargin=1.5*cm
    )

    styles  = getSampleStyleSheet()
    title_s = ParagraphStyle("ttl", parent=styles["Heading2"],
                             alignment=TA_CENTER, spaceAfter=6)
    cell_s  = ParagraphStyle("cel", parent=styles["Normal"],
                             fontSize=7, leading=9, alignment=TA_CENTER)
    free_s  = ParagraphStyle("fre", parent=styles["Normal"],
                             fontSize=7, leading=9, alignment=TA_CENTER,
                             textColor=colors.HexColor("#6C757D"))

    NAVY  = colors.HexColor("#1F3864")
    LBLUE = colors.HexColor("#D9E1F2")
    YFREE = colors.HexColor("#FFF9C4")
    BLAB  = colors.HexColor("#E7F5FF")
    WHITE = colors.white


    # Support single-class export via ?class_idx=N
    single_idx = request.args.get("class_idx", None)
    if single_idx is not None:
        try:
            single_idx = int(single_idx)
            class_indices = [single_idx]
            class_names   = [class_names[single_idx]] if single_idx < len(class_names) else class_names
        except (ValueError, IndexError):
            class_names   = [organized_keys[i] if i < len(organized_keys) else str(i+1) for i in range(num_classes)]
            class_indices = list(range(num_classes))
    else:
        class_indices = list(range(num_classes))

    story = []
    for cls_idx, cls_name in zip(class_indices, class_names):
        story.append(Paragraph(f"Class {cls_name} — Timetable", title_s))

        # Build table rows: header + one row per day
        header = ["Day"] + [f"P{p+1}" for p in range(periods)]
        rows   = [header]

        # Track which (row, col) cells need colour overrides
        free_cells = []
        lab_cells  = []

        for d in range(days):
            row = [Paragraph(day_labels[d], cell_s)]
            for p in range(periods):
                text, kind = _cell_text(timetable, cls_idx, d, p, periods)
                style = free_s if kind == "free" else cell_s
                row.append(Paragraph(text, style))
                if kind == "free":
                    free_cells.append((p+1, d+1))   # col, row
                elif kind == "lab":
                    lab_cells.append((p+1, d+1))
            rows.append(row)

        col_w = (27 * cm) / (periods + 1)
        t = Table(rows, colWidths=[col_w] * (periods + 1), repeatRows=1)

        ts = TableStyle([
            # Header row
            ("BACKGROUND", (0, 0), (-1, 0),  NAVY),
            ("TEXTCOLOR",  (0, 0), (-1, 0),  WHITE),
            ("FONTNAME",   (0, 0), (-1, 0),  "Helvetica-Bold"),
            ("FONTSIZE",   (0, 0), (-1, 0),  9),
            # Day column
            ("BACKGROUND", (0, 1), (0, -1),  LBLUE),
            ("FONTNAME",   (0, 1), (0, -1),  "Helvetica-Bold"),
            # All cells
            ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
            ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
            ("FONTSIZE",   (1, 1), (-1, -1), 8),
            ("ROWHEIGHT",  (0, 1), (-1, -1), 28),
            ("GRID",       (0, 0), (-1, -1), 0.5, colors.grey),
        ])

        # Apply per-cell colour overrides
        for (col_i, row_i) in free_cells:
            ts.add("BACKGROUND", (col_i, row_i), (col_i, row_i), YFREE)
        for (col_i, row_i) in lab_cells:
            ts.add("BACKGROUND", (col_i, row_i), (col_i, row_i), BLAB)

        t.setStyle(ts)
        story.append(t)
        story.append(Spacer(1, 0.8 * cm))

    doc.build(story)
    output.seek(0)
    fname = f"timetable_class_{class_names[0]}.pdf" if len(class_names) == 1 else "timetable.pdf"
    return send_file(
        output,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=fname
    )















# CLEANED: Only one version of success_summary using dynamic metadata
@app.route("/success-summary")
def success_summary():
    if not os.path.exists("generated_timetable.json") or not os.path.exists("generated_metadata.json"):
        return redirect(url_for('home'))

    with open("generated_timetable.json", "r") as f:
        timetable = json.load(f)
    with open("generated_metadata.json", "r") as f:
        meta = json.load(f)
    with open("final_schedule.json", "r") as f:
        final_data = json.load(f)
    with open("temp_web_data.json", "r") as f:
        stored = json.load(f)

    days        = meta['days']
    periods     = meta['periods']
    num_classes = meta['num_classes']

    class_names_raw = list(dict.fromkeys([row['class'] for row in final_data]))
    # Strip "Class " prefix stored in final_schedule
    class_names = [c.replace("Class ", "").strip() for c in class_names_raw]

    # ── Build teacher_slot_map: {teacher_name: ["classIdx-slotIdx", ...]} ──────
    # We need to know which teacher teaches each subject in each class
    organized = stored.get('organized', {})
    # subject->teacher lookup per class
    subj_teacher = {}  # (class_idx, subject_lower) -> teacher_name
    for cidx, cname in enumerate(organized.keys()):
        for t in organized[cname]:
            key = (cidx, t['subject'].lower().strip())
            subj_teacher[key] = t['teacher']

    teacher_slot_map = {}   # teacher_name -> [classIdx-slotIdx]
    teacher_names_set = set()
    for cidx in range(num_classes):
        for day in range(days):
            for p in range(periods):
                si = day * periods + p
                try:
                    cell = timetable[si][cidx]
                except (IndexError, KeyError):
                    continue
                if not cell or cell == 0 or str(cell).lower() in ('free', 'f', '0'):
                    continue
                cell_str = str(cell).strip()
                key = (cidx, cell_str.lower().strip())
                tname = subj_teacher.get(key)
                # Fuzzy match: check if cell_str starts with any known subject
                if not tname:
                    for (c2, subj), tn in subj_teacher.items():
                        if c2 == cidx and cell_str.lower().startswith(subj[:6]):
                            tname = tn
                            break
                if tname:
                    teacher_names_set.add(tname)
                    if tname not in teacher_slot_map:
                        teacher_slot_map[tname] = []
                    teacher_slot_map[tname].append(f"{cidx}-{si}")

    teacher_names = sorted(teacher_names_set)

    # ── Conflict checker: same teacher in 2 classes at same slot ─────────────
    conflicts = []
    slot_teacher_classes = {}  # (tname, si) -> [class_idxs]
    for tname, slots in teacher_slot_map.items():
        for s in slots:
            cidx_str, si_str = s.split('-')
            key = (tname, int(si_str))
            slot_teacher_classes.setdefault(key, []).append(int(cidx_str))
    for (tname, si), cidxs in slot_teacher_classes.items():
        if len(cidxs) > 1:
            for cidx in cidxs:
                conflicts.append([cidx, si])

    # teacher_map: {teacher_name: teacher_id} for JS
    teacher_map_js = {t['teacher']: t['teacher_id']
                      for cname in organized for t in organized[cname]}

    return render_template("success.html",
                           timetable=timetable,
                           num_classes=num_classes,
                           class_names=class_names,
                           num_days=days,
                           periods_per_day=periods,
                           teacher_names=teacher_names,
                           teacher_slot_map=teacher_slot_map,
                           teacher_map=teacher_map_js,
                           conflicts=conflicts)





# --- KEEP THIS VERSION (REPLACES THE TWO OLD ONES) ---
@app.route("/update-data", methods=["POST"])
def update_data():
    try:
        incoming_payload = request.get_json()
        web_data = incoming_payload.get('table_data', [])
        config = incoming_payload.get('config', {})
        
        # Create a mapping of teacher names to unique IDs
        all_teachers = sorted(list(set(row['teacher'] for row in web_data)))
        t_name_to_id = {name: i for i, name in enumerate(all_teachers)}

        organized_classes = {}
        # Inside @app.route("/update-data", methods=["POST"])
        for row in web_data:
            c_name = row['class'].replace("Class ", "").strip()
            if c_name not in organized_classes: 
                organized_classes[c_name] = []
            
            # ENSURE THESE TYPES ARE EXPLICIT
            organized_classes[c_name].append({
                "teacher": row.get('teacher', 'Unknown'),
                "teacher_id": t_name_to_id.get(row.get('teacher'), 99), 
                "subject": row.get('subject', 'General'),
                "hours": int(row.get('periods', 0)),
                "type": str(row.get('type', 'theory')).lower().strip(), # Clean string
                "continuous": int(row.get('continuous', 1)),
                "lab_no": int(row.get('lab_no', 0))
            })

        session_data = {
            "organized": organized_classes,
            "days": int(config.get('days', 6)),
            "periods": int(config.get('periods', 6)),
            "session_token": str(__import__('uuid').uuid4())  # fresh token every upload
        }
        with open("temp_web_data.json", "w") as f:
            json.dump(session_data, f)

        return jsonify({"status": "success", "redirect": url_for('setup_fixed')})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


# .................
@app.route("/setup-fixed")
def setup_fixed():
    if not os.path.exists("temp_web_data.json"):
        return redirect(url_for('home'))
        
    with open("temp_web_data.json", "r") as f:
        stored = json.load(f)
    
    # Defensive: always provide periods
    periods_value = stored.get('periods', 8)          # fallback to 8 if missing
    if not isinstance(periods_value, (int, float)):
        periods_value = 8
    
    return render_template(
        "fixed_setup.html",
        days=stored.get('days', 6),
        periods=periods_value,
        class_data=stored.get('organized', {}),
        session_token=stored.get('session_token', 'default')
    )
@app.route("/run-final-solver", methods=["POST"])
def run_final_solver():
    try:
        payload        = request.get_json()
        fixed_data     = payload.get('fixed_slots', {})
        unavail_data   = payload.get('teacher_unavailability', {})
        
        if not os.path.exists("temp_web_data.json"):
            return jsonify({"status": "error", "message": "Session expired. Please restart."}), 400

        with open("temp_web_data.json", "r") as f:
            stored = json.load(f)

        from adapter import build_final_inputs 
        
        (No_of_classes, t_list, c_theory, l_periods, subj_map) = build_final_inputs(
            {"classes": stored['organized']}, 
            stored['days'], 
            stored['periods'], 
            fixed_data
        )

        # --- DEBUG LOGGING: Save the exact data going into the solver ---
        debug_payload = {
            "No_of_classes": No_of_classes,
            "days": stored['days'],
            "periods": stored['periods'],
            "teacher_list": t_list,
            "class_theory_workload": c_theory,
            "lab_periods": l_periods,
            "subject_map": {str(k): v for k, v in subj_map.items()},  # Convert tuple keys to strings for JSON
            "fixed_periods": fixed_data
        }
        with open("solver_input_debug.json", "w") as f:
            json.dump(debug_payload, f, indent=4)
        # --------------------------------------------------------------

        final_timetable = generate_timetable_with_retry(
            No_of_classes, stored['days'], stored['periods'], t_list,
            c_theory, l_periods, subj_map,
            fixed_periods=fixed_data,
            teacher_unavailability=unavail_data
        )

        if final_timetable:
            # 1. Save metadata for the success page
            with open("generated_metadata.json", "w") as f:
                json.dump({
                    "days": stored['days'], 
                    "periods": stored['periods'], 
                    "num_classes": No_of_classes
                }, f)
            
            # 2. Save the actual timetable
            with open("generated_timetable.json", "w") as f:
                json.dump(final_timetable, f)

            # 3. Create the 'final_schedule.json' that success_summary expects
            flat_rows = []
            for c_name, teachers in stored['organized'].items():
                for t in teachers:
                    flat_rows.append({"class": f"Class {c_name}", "teacher": t['teacher']})
            
            with open("final_schedule.json", "w") as f:
                json.dump(flat_rows, f)

            return jsonify({"status": "success", "redirect": url_for('success_summary')})
        
        # Build a conflict report by checking teacher workload vs available slots
        report_lines = []
        for cidx, cname in enumerate(stored['organized'].keys()):
            total_slots = stored['days'] * stored['periods']
            teachers = stored['organized'][cname]
            total_hours = sum(int(t.get('hours', 0)) for t in teachers)
            if total_hours > total_slots:
                report_lines.append(f"Class {cname}: {total_hours} hours assigned but only {total_slots} slots available.")
        # Check teacher double-booking across classes
        teacher_class_hours = {}
        for cname, teachers in stored['organized'].items():
            for t in teachers:
                tid = t.get('teacher_id')
                tname = t.get('teacher', f'T{tid}')
                hours = int(t.get('hours', 0))
                teacher_class_hours.setdefault(tname, 0)
                teacher_class_hours[tname] += hours
        max_slots = stored['days'] * stored['periods']
        for tname, total in teacher_class_hours.items():
            if total > max_slots:
                report_lines.append(f"Teacher <b>{tname}</b> has {total} total hours across all classes but only {max_slots} slots/week.")
        conflict_report = "<br>".join(report_lines) if report_lines else "No specific conflict identified. Try removing some fixed slots or unavailability constraints."
        return jsonify({"status": "error", "message": "Solver failed.", "conflict_report": conflict_report})
    except Exception as e:
        import traceback
        print(traceback.format_exc()) 
        return jsonify({"status": "error", "message": str(e)}), 500

        


@app.route("/swap-slots", methods=["POST"])
def swap_slots():
    try:
        data      = request.get_json()
        class_idx = int(data['class_idx'])
        si1       = int(data['slot1'])
        si2       = int(data['slot2'])

        if not os.path.exists("generated_timetable.json"):
            return jsonify({"status": "error", "message": "No timetable found"}), 404

        with open("generated_timetable.json") as f:
            timetable = json.load(f)

        # Swap the two slots for the given class
        timetable[si1][class_idx], timetable[si2][class_idx] = \
            timetable[si2][class_idx], timetable[si1][class_idx]

        with open("generated_timetable.json", "w") as f:
            json.dump(timetable, f)

        return jsonify({"status": "success"})
    except Exception as e:
        import traceback; print(traceback.format_exc())
        return jsonify({"status": "error", "message": str(e)}), 500

if __name__ == "__main__":
    app.run(host='0.0.0.0', port=5000, debug=True)